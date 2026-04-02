from __future__ import annotations

import csv
import json
import re
import sys
import threading
from datetime import datetime
from pathlib import Path
from typing import Callable
from urllib.parse import urlparse

import requests
from charset_normalizer import from_bytes
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from PyQt5.QtCore import Qt, QThread, QUrl, pyqtSignal
from PyQt5.QtGui import QDesktopServices, QFont, QCloseEvent
from PyQt5.QtWidgets import (
    QApplication,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPlainTextEdit,
    QPushButton,
    QVBoxLayout,
    QWidget,
)

OnLog = Callable[[str], None]
OnProgress = Callable[[int, int], None]
OnMaterialRoot = Callable[[str], None]

CSV_PROBE_BYTES = 256 * 1024

FALLBACK_ENCODINGS: tuple[str, ...] = (
    "gb18030",
    "gbk",
    "cp936",
    "big5",
    "cp950",
    "utf-16-le",
    "utf-16-be",
    "utf-16",
)

PREFERRED_UTF: tuple[str, ...] = ("utf-8-sig", "utf-8")


def _normalize_enc_key(name: str) -> str:
    return name.replace("_", "-").lower()


def build_csv_encoding_candidates(sample: bytes) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()

    def add(enc: str | None) -> None:
        if not enc:
            return
        key = _normalize_enc_key(enc)
        if key in seen:
            return
        seen.add(key)
        ordered.append(enc)

    for enc in PREFERRED_UTF:
        add(enc)

    try:
        matches = from_bytes(sample)
        for i, m in enumerate(matches):
            if i >= 6:
                break
            add(m.encoding)
    except Exception:
        pass

    for enc in FALLBACK_ENCODINGS:
        add(enc)

    return ordered


def load_csv_with_encoding_detection(
    csv_path: Path,
    on_log: OnLog | None = None,
) -> tuple[list[str], list[dict[str, str]], str]:
    size = csv_path.stat().st_size
    if size == 0:
        raise ValueError("CSV 文件为空。")
    sample = csv_path.read_bytes()[: min(CSV_PROBE_BYTES, size)]
    candidates = build_csv_encoding_candidates(sample)
    if on_log:
        on_log(f"CSV 编码候选（探测 + 回退）: {', '.join(candidates[:8])}{'…' if len(candidates) > 8 else ''}")

    errors: list[str] = []
    for enc in candidates:
        try:
            with open(csv_path, "r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    errors.append(f"{enc}: 无表头")
                    continue
                headers = list(reader.fieldnames)
                rows = list(reader)
            return headers, rows, enc
        except UnicodeDecodeError as e:
            errors.append(f"{enc}: {e}")
        except UnicodeError as e:
            errors.append(f"{enc}: {e}")

    detail = "; ".join(errors[:5])
    if len(errors) > 5:
        detail += "…"
    raise ValueError(f"无法用已知编码解码 CSV。尝试摘要: {detail}")


WIN_ILLEGAL = r'<>:"/\\|?*'
WIN_ILLEGAL_SET = set(WIN_ILLEGAL)

SESSION_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "*/*",
}

IMG_SRC_RE = re.compile(
    r'<img[^>]+src\s*=\s*(["\'])(.*?)\1',
    re.IGNORECASE | re.DOTALL,
)

PIPE_DELIM = "'|'"

CONTENT_TYPE_EXT: dict[str, str] = {
    "image/jpeg": ".jpg",
    "image/jpg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "image/gif": ".gif",
    "image/bmp": ".bmp",
    "image/heic": ".heic",
    "image/heif": ".heif",
    "video/mp4": ".mp4",
    "video/quicktime": ".mov",
    "video/webm": ".webm",
    "video/x-msvideo": ".avi",
}

VIDEO_EXTS = {".mp4", ".mov", ".webm", ".avi", ".mkv", ".m4v", ".wmv"}
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp", ".heic", ".heif"}

MAX_PRODUCT_FOLDER_LEN = 120
MAX_STEM_SANITIZE = 200
MAX_SKU_NAME_STEM = 80


def sanitize_component(text: str, max_len: int) -> str:
    if not text:
        return ""
    s = "".join("_" if c in WIN_ILLEGAL_SET or ord(c) < 32 else c for c in str(text).strip())
    s = s.rstrip(" .")
    if not s:
        s = "_"
    if len(s) > max_len:
        s = s[:max_len].rstrip(" .") or "_"
    upper = s.upper()
    if upper in {"CON", "PRN", "AUX", "NUL", "COM1", "COM2", "COM3", "COM4", "LPT1", "LPT2", "LPT3"}:
        s = f"_{s}"
    return s


def product_folder_name(name: str, product_id: str) -> str:
    id_part = sanitize_component(product_id, 40)
    budget = MAX_PRODUCT_FOLDER_LEN - len(id_part) - 1
    if budget < 1:
        id_part = id_part[:20]
        budget = MAX_PRODUCT_FOLDER_LEN - len(id_part) - 1
    name_part = sanitize_component(name, min(budget, MAX_PRODUCT_FOLDER_LEN))
    out = f"{name_part}_{id_part}"
    if len(out) > MAX_PRODUCT_FOLDER_LEN:
        out = out[:MAX_PRODUCT_FOLDER_LEN].rstrip("._ ") or "_"
    return out


def make_unique_product_dir(
    root_dir: Path,
    name: str,
    product_id: str,
    on_log: OnLog,
    idx: int,
    total: int,
) -> Path | None:
    """在 root_dir 下创建商品文件夹；若与已有目录重名则自动追加 _2、_3 …"""
    base_folder = product_folder_name(name or "_", product_id)
    folder = base_folder
    counter = 1
    max_attempts = 10_000

    while True:
        product_dir = root_dir / folder
        try:
            product_dir.mkdir(parents=False, exist_ok=False)
            if folder != base_folder:
                on_log(f"[{idx}/{total}] 文件夹名与已有记录冲突，已改用：{folder}")
            return product_dir
        except FileExistsError:
            counter += 1
            if counter > max_attempts:
                on_log(
                    f"[{idx}/{total}] 跳过：无法为「{base_folder}」分配唯一文件夹（已达 {max_attempts} 次）。"
                )
                return None
            suffix = f"_{counter}"
            max_base = MAX_PRODUCT_FOLDER_LEN - len(suffix)
            if max_base < 1:
                max_base = 1
            trimmed = base_folder[:max_base].rstrip("._ ") or "_"
            folder = trimmed + suffix
        except OSError as e:
            on_log(f"[{idx}/{total}] 跳过：无法创建文件夹 {folder} — {e}")
            return None


def root_material_folder_name(csv_stem: str) -> str:
    stem = sanitize_component(csv_stem, MAX_STEM_SANITIZE)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{stem}素材包{ts}"


def split_pipe_field(value: str | None) -> list[str]:
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    parts = [p.strip() for p in s.split(PIPE_DELIM)]
    return [p for p in parts if p]


def row_has_product_id(row: dict[str, str]) -> bool:
    return bool((row.get("商品ID") or "").strip())


def get_valid_sku_indices(row: dict[str, str]) -> list[int]:
    """过滤 SKU属性 中与商品名称完全相等（去首尾空格后）的脏元素。"""
    product_name = (row.get("商品名称") or "").strip()
    skus = split_pipe_field(row.get("SKU属性"))
    out: list[int] = []
    for i, sku in enumerate(skus):
        if sku.strip() == product_name:
            continue
        out.append(i)
    return out


def row_has_skus(row: dict[str, str]) -> bool:
    return len(get_valid_sku_indices(row)) > 0


def get_csv_data_row_count(csv_path: Path) -> tuple[int | None, str | None]:
    """统计有效数据行数：商品ID 非空且 SKU属性 至少有一段；不含表头。"""
    p = csv_path.resolve()
    if not p.is_file():
        return None, "文件不存在。"
    try:
        headers, rows, _enc = load_csv_with_encoding_detection(p, None)
    except ValueError as e:
        return None, str(e)
    if "商品ID" not in headers or "商品名称" not in headers:
        return None, "CSV 缺少必需列：商品ID 或 商品名称。"
    return sum(1 for r in rows if row_has_product_id(r) and row_has_skus(r)), None


PRODUCT_INFO_HEADERS: tuple[str, ...] = (
    "*标题",
    "货号",
    "商品属性",
    "类目",
    "品牌",
    "规格1",
    "规格2",
    "*价格",
    "库存",
    "短标题",
    "商家SKU",
    "SKU商品条形码",
    "SKU属性",
    "无理由退货",
    "支付方式限制",
    "产地",
    "发货地",
    "商品条形码",
    "商品毛重(公斤)",
    "[包装]长(mm)",
    "[包装]宽(mm)",
    "[包装]高(mm)",
)

TPL_COL_TITLE = 1
TPL_COL_HUOHAO = 2
TPL_COL_ATTR = 3
TPL_COL_CATEGORY = 4
TPL_COL_BRAND = 5
TPL_COL_SPEC1 = 6
TPL_COL_SPEC2 = 7
TPL_COL_PRICE = 8
TPL_COL_STOCK = 9
TPL_COL_SHORT_TITLE = 10
TPL_COL_MERCHANT_SKU = 11
TPL_COL_SKU_BARCODE = 12


def format_product_attr_cell(raw: str | None) -> str:
    s = (raw or "").strip()
    if not s:
        return ""
    try:
        obj = json.loads(s)
    except json.JSONDecodeError:
        return s
    if not isinstance(obj, dict):
        return s
    parts = [f"{k}: {str(v)}" for k, v in obj.items()]
    return ", ".join(parts)


def format_category_for_template(raw: str | None) -> str:
    s = (raw or "").strip()
    if not s:
        return ""
    return s.replace("-->", ">")


def split_sku_spec(raw: str) -> tuple[str, str]:
    s = (raw or "").strip()
    if not s:
        return "", ""
    parts = [p.strip() for p in s.split(";") if p.strip()]
    if len(parts) >= 2:
        return parts[1], parts[0]
    return s, ""


def build_short_title(spec1: str, spec2: str, max_len: int = 15) -> str:
    s = f"{(spec1 or '').strip()}{(spec2 or '').strip()}".strip()
    if not s:
        return "_"
    if len(s) > max_len:
        return s[:max_len]
    return s


def _maybe_excel_number(s: str) -> str | int | float:
    t = (s or "").strip()
    if not t:
        return ""
    try:
        if "." in t:
            return float(t)
        return int(t)
    except ValueError:
        return t


TEMPLATE_COL_WIDTHS: dict[int, float] = {
    1: 44.8416666666667,
    2: 8.875,
    3: 36.0,
    4: 35.0,
    5: 18.0333333333333,
    6: 13.875,
    7: 15.625,
    8: 12.875,
    9: 9.0,
    10: 11.6,
    11: 9.875,
    12: 16.425,
    13: 36.0666666666667,
    14: 17.75,
    15: 16.0,
    16: 17.9583333333333,
    17: 13.25,
    18: 21.8083333333333,
    19: 16.625,
    20: 13.0,
    21: 13.0,
    22: 13.0,
}


def setup_product_sheet(ws) -> int:
    ws.title = "商品信息"

    for col, width in TEMPLATE_COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 134.0
    ws.row_dimensions[2].height = 17.25

    ws.merge_cells("A1:D1")
    ws["A1"].value = (
        "注意事项\n"
        "1.注意：某些平台规则较严，请不要填写违规商品，不要填写重复信息。\n"
        "2.同一商品多条SKU，SKU信息（需要填的价格、库存）按行填写；SPU信息只需填写一行即可（如类目）。\n"
        "3.主图、SKU图、详情图通过上传图片的方式上传，素材包上传格式请看“素材包使用教程.docx”。\n"
        "4.带*字段为必填项，漏填将上传失败。\n"
        "5.类目和发货地需要按平台的格式填写，否则无法匹配（可在商家端过程中编辑）。\n"
        "6.[包装]字段默认单位为毫米，可为小数。\n"
        "7.属性格式：属性名:属性值；多个值用逗号分隔。"
    )
    ws["A1"].font = Font(name="微软雅黑", size=11, bold=True, color="FFFF0000")
    ws["A1"].alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )

    header_font = Font(name="微软雅黑", size=12, color="FF000000")
    header_font_required = Font(name="微软雅黑", size=12, color="FFFF0000")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, title in enumerate(PRODUCT_INFO_HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=title)
        cell.font = header_font_required if str(title).startswith("*") else header_font
        cell.alignment = header_alignment

    # 条码类字段按文本处理，避免科学计数法/丢前导 0
    ws.cell(row=2, column=TPL_COL_SKU_BARCODE).number_format = "@"
    ws.cell(row=2, column=18).number_format = "@"

    return 3


def _row_has_spec2(row: dict[str, str], on_log: OnLog | None = None) -> bool:
    skus = split_pipe_field(row.get("SKU属性"))
    valid_indices = get_valid_sku_indices(row)
    for i in valid_indices:
        if i >= len(skus):
            continue
        spec1, spec2 = split_sku_spec(skus[i])
        if spec2:
            if on_log is not None:
                pid = (row.get("商品ID") or "").strip()
                name = (row.get("商品名称") or "").strip()
                on_log(
                    "检测到规格2（按商品判断）："
                    f"商品ID={pid or '_'}，商品名称={name or '_'}，"
                    f"SKU段={skus[i]!r} -> 规格1={spec1!r} 规格2={spec2!r}"
                )
            return True
    return False


def write_product_xlsx(
    dest: Path,
    row: dict[str, str],
    *,
    has_spec2: bool,
) -> None:
    wb = Workbook()
    ws = wb.active
    first_data_row = setup_product_sheet(ws)

    skus = split_pipe_field(row.get("SKU属性"))
    prices = split_pipe_field(row.get("SKU价格"))
    stocks = split_pipe_field(row.get("SKU库存"))
    barcodes = split_pipe_field(row.get("SKU条形码"))
    merchant_skus = split_pipe_field(row.get("SKU 商家编码"))

    title = (row.get("商品名称") or "").strip()
    huohao = (row.get("商品ID") or "").strip()
    attr_text = format_product_attr_cell(row.get("商品属性"))
    category = format_category_for_template(row.get("商品分类"))
    brand = "无"

    valid_indices = get_valid_sku_indices(row)
    # 按规格1排序写入，确保同规格1相邻；排序对象为索引，保证整行字段关联不变
    sorted_indices = sorted(
        valid_indices,
        key=lambda i: split_sku_spec(skus[i])[0] if i < len(skus) else "",
    )
    for out_i, src_i in enumerate(sorted_indices):
        r = first_data_row + out_i
        ws.row_dimensions[r].height = 45.0

        spec1, spec2 = split_sku_spec(skus[src_i])
        short_title = build_short_title(spec1, spec2)
        price_s = prices[src_i] if src_i < len(prices) else ""
        stock_s = stocks[src_i] if src_i < len(stocks) else ""
        bc_s = barcodes[src_i] if src_i < len(barcodes) else ""
        msku_s = merchant_skus[src_i] if src_i < len(merchant_skus) else ""

        if out_i == 0:
            ws.cell(row=r, column=TPL_COL_TITLE, value=title)
            ws.cell(row=r, column=TPL_COL_HUOHAO, value=huohao)
            ws.cell(row=r, column=TPL_COL_ATTR, value=attr_text)
            ws.cell(row=r, column=TPL_COL_CATEGORY, value=category)
            ws.cell(row=r, column=TPL_COL_BRAND, value=brand)
        else:
            for c in range(1, TPL_COL_BRAND + 1):
                ws.cell(row=r, column=c, value=None)

        ws.cell(row=r, column=TPL_COL_SPEC1, value=spec1)
        ws.cell(
            row=r,
            column=TPL_COL_SPEC2,
            value=spec2 if spec2 else ("*" if has_spec2 else None),
        )
        ws.cell(row=r, column=TPL_COL_PRICE, value=price_s or None)
        ws.cell(row=r, column=TPL_COL_STOCK, value=stock_s or None)
        ws.cell(row=r, column=TPL_COL_SHORT_TITLE, value=short_title)
        ws.cell(row=r, column=TPL_COL_MERCHANT_SKU, value=msku_s or None)
        ws.cell(row=r, column=TPL_COL_SKU_BARCODE, value=bc_s or None)
        ws.cell(row=r, column=TPL_COL_SKU_BARCODE).number_format = "@"

    wb.save(dest)


def extension_from_url(url: str) -> str:
    try:
        suf = Path(urlparse(url).path).suffix.lower()
    except Exception:
        return ""
    if suf in IMAGE_EXTS or suf in VIDEO_EXTS:
        return suf
    if suf == ".jpe":
        return ".jpg"
    return ""


def extension_from_content_type(ct: str | None) -> str:
    if not ct:
        return ""
    base = ct.split(";")[0].strip().lower()
    return CONTENT_TYPE_EXT.get(base, "")


def pick_image_extension(url: str, content_type: str | None) -> str:
    u = extension_from_url(url)
    if u:
        return u
    c = extension_from_content_type(content_type)
    if c:
        return c
    return ".jpg"


def pick_video_extension(url: str, content_type: str | None) -> str:
    u = extension_from_url(url)
    if u and u in VIDEO_EXTS:
        return u
    c = extension_from_content_type(content_type)
    if c and c in VIDEO_EXTS:
        return c
    if u:
        return u
    return ".mp4"


def extract_img_srcs(html: str | None) -> list[str]:
    if not html:
        return []
    out: list[str] = []
    for m in IMG_SRC_RE.finditer(html):
        src = (m.group(2) or "").strip()
        if src.lower().startswith("http://") or src.lower().startswith("https://"):
            out.append(src)
    return out


def sanitize_file_stem(name: str, max_len: int = MAX_SKU_NAME_STEM) -> str:
    s = sanitize_component(name, max_len)
    s = s.replace(" ", "_")
    return s


def parse_sku_image_base_name(raw: str) -> str:
    """从 SKU属性 子元素中提取 SKU图基础名。"""
    s = (raw or "").strip()
    if not s:
        return ""
    semicolon_count = s.count(";")
    if semicolon_count == 1:
        parts = [p.strip() for p in s.split(";")]
        if len(parts) >= 2 and parts[1]:
            return parts[1]
    # 无分号或分号超过 1 个时，不做二次分割，直接使用原值
    return s


def download_to_file(
    url: str,
    dest_without_ext: Path,
    stop_event: threading.Event,
    on_log: OnLog,
    *,
    is_video: bool,
) -> bool:
    if stop_event.is_set():
        return False
    last_err: str | None = None
    session = requests.Session()
    session.trust_env = False
    for attempt in range(1, 4):
        if stop_event.is_set():
            return False
        try:
            with session.get(
                url,
                timeout=30,
                headers=SESSION_HEADERS,
                stream=True,
            ) as resp:
                resp.raise_for_status()
                ct = resp.headers.get("Content-Type")
                ext = (
                    pick_video_extension(url, ct)
                    if is_video
                    else pick_image_extension(url, ct)
                )
                final_path = dest_without_ext.with_suffix(ext)
                tmp_path = final_path.with_name(final_path.name + ".part")
                with open(tmp_path, "wb") as f:
                    for chunk in resp.iter_content(chunk_size=65536):
                        if chunk:
                            f.write(chunk)
                if final_path.exists():
                    final_path.unlink()
                tmp_path.replace(final_path)
            return True
        except Exception as e:
            last_err = str(e)
            on_log(f"下载失败（第 {attempt} 次）: {url[:80]}… — {e}")
    if last_err:
        on_log(f"已跳过（3 次均失败）: {url[:80]}…")
    return False


def run_job(
    csv_path: Path,
    output_base: Path,
    stop_event: threading.Event,
    on_log: OnLog,
    on_progress: OnProgress | None = None,
    on_material_root: OnMaterialRoot | None = None,
) -> tuple[bool, str, str]:
    csv_path = csv_path.resolve()
    output_base = output_base.resolve()
    if not csv_path.is_file():
        return False, "CSV 文件不存在。", ""
    if not output_base.is_dir():
        return False, "输出目录无效。", ""

    try:
        headers, rows, used_enc = load_csv_with_encoding_detection(csv_path, on_log)
    except ValueError as e:
        return False, str(e), ""

    on_log(f"已使用编码读取 CSV: {used_enc}")

    if "商品ID" not in headers or "商品名称" not in headers:
        return False, "CSV 缺少必需列：商品ID 或 商品名称。", ""

    main_cols = [f"主图{i}" for i in range(1, 6)]

    root_name = root_material_folder_name(csv_path.stem)
    root_dir = output_base / root_name
    try:
        root_dir.mkdir(parents=True, exist_ok=False)
    except FileExistsError:
        return False, f"素材包目录已存在（可能同一秒内重复运行）: {root_dir}", ""
    except OSError as e:
        return False, f"无法创建素材包目录: {e}", ""

    material_root_str = str(root_dir)
    on_log(f"素材包目录: {root_dir}")
    if on_material_root is not None:
        on_material_root(material_root_str)

    raw_row_count = len(rows)
    rows_with_id = [r for r in rows if row_has_product_id(r)]
    skipped_empty_id = raw_row_count - len(rows_with_id)
    effective_rows = [r for r in rows_with_id if row_has_skus(r)]
    skipped_no_sku = len(rows_with_id) - len(effective_rows)
    total = len(effective_rows)
    parts_summary: list[str] = []
    if skipped_empty_id:
        parts_summary.append(f"{skipped_empty_id} 行商品ID 为空已忽略")
    if skipped_no_sku:
        parts_summary.append(f"{skipped_no_sku} 行 SKU属性 无有效分段已忽略")
    if parts_summary:
        on_log(
            f"CSV 共 {raw_row_count} 行数据，{', '.join(parts_summary)}，"
            f"待处理 {total} 条，开始处理。"
        )
    else:
        on_log(f"共 {total} 条有效商品记录，开始处理。")

    if total == 0:
        on_log("没有商品ID 非空的记录，无需处理。")
        on_log("全部商品处理完成。")
        return True, "完成", material_root_str

    for idx, row in enumerate(effective_rows, start=1):
        if stop_event.is_set():
            on_log("已收到停止指令，结束任务（当前文件已写完）。")
            return True, "已停止", material_root_str

        if on_progress is not None:
            on_progress(idx, total)

        name = (row.get("商品名称") or "").strip()
        pid = (row.get("商品ID") or "").strip()

        product_dir = make_unique_product_dir(root_dir, name, pid, on_log, idx, total)
        if product_dir is None:
            continue
        folder = product_dir.name

        sub_main = product_dir / "主图"
        sub_sku = product_dir / "SKU图"
        sub_detail = product_dir / "详情图"
        sub_video = product_dir / "主图视频"
        for d in (sub_main, sub_sku, sub_detail, sub_video):
            d.mkdir(parents=True, exist_ok=True)

        if stop_event.is_set():
            on_log("已收到停止指令，结束任务。")
            return True, "已停止", material_root_str

        xlsx_path = product_dir / "商品信息.xlsx"
        try:
            has_spec2 = _row_has_spec2(row, on_log)
            write_product_xlsx(xlsx_path, row, has_spec2=has_spec2)
            on_log(f"[{idx}/{total}] 已写入 商品信息.xlsx（{folder}）")
        except Exception as e:
            on_log(f"[{idx}/{total}] 写入 xlsx 失败: {e}")

        for mi, col in enumerate(main_cols, start=1):
            if col not in row:
                continue
            raw = (row.get(col) or "").strip()
            if not raw:
                continue
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str
            dest_stem = sub_main / f"主图_{mi}"
            on_log(f"[{idx}/{total}] 下载 主图_{mi} …")
            ok = download_to_file(raw, dest_stem, stop_event, on_log, is_video=False)
            if ok:
                on_log(f"[{idx}/{total}] 主图_{mi} 完成")
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str

        desc = row.get("商品描述图")
        imgs = extract_img_srcs(desc)
        for n, u in enumerate(imgs, start=1):
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str
            stem = sub_detail / f"详情图_{n}"
            on_log(f"[{idx}/{total}] 下载 详情图_{n} …")
            ok = download_to_file(u, stem, stop_event, on_log, is_video=False)
            if ok:
                on_log(f"[{idx}/{total}] 详情图_{n} 完成")
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str

        urls_sku = split_pipe_field(row.get("SKU属性图"))
        names_sku = split_pipe_field(row.get("SKU属性"))
        valid_sku_indices = get_valid_sku_indices(row)
        filtered_sku_names = [
            names_sku[i] for i in valid_sku_indices if i < len(names_sku)
        ]
        filtered_sku_urls = [
            urls_sku[i] for i in valid_sku_indices if i < len(urls_sku)
        ]
        if len(urls_sku) != len(names_sku):
            on_log(
                f"[{idx}/{total}] 提示：SKU属性图（{len(urls_sku)} 个）与 "
                f"SKU属性（{len(names_sku)} 个）数量不一致，缺名将使用 SKU图_N。"
            )
        dropped_dirty = len(names_sku) - len(filtered_sku_names)
        if dropped_dirty > 0:
            on_log(
                f"[{idx}/{total}] 已过滤 SKU属性 脏数据 {dropped_dirty} 个（与商品名称相同）。"
            )
        sku_name_counter: dict[str, int] = {}
        for si, u in enumerate(filtered_sku_urls, start=1):
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str
            if si <= len(filtered_sku_names) and filtered_sku_names[si - 1]:
                base_raw = parse_sku_image_base_name(filtered_sku_names[si - 1])
                base_name = sanitize_file_stem(base_raw)
                if not base_name:
                    base_name = "SKU图"
            else:
                base_name = "SKU图"
            next_no = sku_name_counter.get(base_name, 0) + 1
            sku_name_counter[base_name] = next_no
            stem_name = f"{base_name}_{next_no}"
            dest_stem = sub_sku / stem_name
            on_log(f"[{idx}/{total}] 下载 SKU 图 ({stem_name}) …")
            ok = download_to_file(u, dest_stem, stop_event, on_log, is_video=False)
            if ok:
                on_log(f"[{idx}/{total}] SKU 图 {stem_name} 完成")
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str

        vid = (row.get("商品视频") or "").strip()
        if vid:
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str
            dest_stem = sub_video / "主图视频"
            on_log(f"[{idx}/{total}] 下载 商品视频 …")
            ok = download_to_file(vid, dest_stem, stop_event, on_log, is_video=True)
            if ok:
                on_log(f"[{idx}/{total}] 主图视频 完成")
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str

        on_log(f"[{idx}/{total}] 商品处理完成：{folder}")

    on_log("全部商品处理完成。")
    return True, "完成", material_root_str


def open_local_dir(path: Path) -> bool:
    p = path.resolve()
    if not p.is_dir():
        return False
    return QDesktopServices.openUrl(QUrl.fromLocalFile(str(p)))


class PutawayThread(QThread):
    log_message = pyqtSignal(str)
    progress = pyqtSignal(int, int)
    material_root_created = pyqtSignal(str)
    finished = pyqtSignal(bool, str, str)

    def __init__(self, csv_path: Path, output_dir: Path, stop_event: threading.Event) -> None:
        super().__init__()
        self._csv_path = csv_path
        self._output_dir = output_dir
        self._stop_event = stop_event

    def run(self) -> None:
        def emit_log(text: str) -> None:
            self.log_message.emit(text)

        def emit_progress(current: int, total: int) -> None:
            self.progress.emit(current, total)

        def emit_material_root(path: str) -> None:
            self.material_root_created.emit(path)

        try:
            ok, msg, material_root = run_job(
                self._csv_path,
                self._output_dir,
                self._stop_event,
                emit_log,
                emit_progress,
                emit_material_root,
            )
            self.finished.emit(ok, msg, material_root)
        except Exception as e:
            self.log_message.emit(f"处理异常: {e}")
            self.finished.emit(False, str(e), "")


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("自动上架工具")
        self.setMinimumSize(720, 560)
        self._csv_path: Path | None = None
        self._total_rows: int | None = None
        self._is_running = False
        self._stop_event: threading.Event | None = None
        self._worker: PutawayThread | None = None
        self._last_material_root: Path | None = None

        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)

        file_group = QGroupBox("表格文件（本地 CSV）")
        file_layout = QHBoxLayout(file_group)
        self.csv_path_edit = QLineEdit()
        self.csv_path_edit.setReadOnly(True)
        self.csv_path_edit.setPlaceholderText("请选择 CSV 文件…")
        self.btn_pick_csv = QPushButton("选择文件…")
        self.btn_pick_csv.clicked.connect(self._on_pick_csv)
        file_layout.addWidget(self.csv_path_edit, 1)
        file_layout.addWidget(self.btn_pick_csv)
        root.addWidget(file_group)

        stat_layout = QHBoxLayout()
        self.lbl_pending = QLabel("待处理：— 条（商品ID 非空且含 SKU）")
        self.lbl_progress = QLabel("当前处理：—")
        stat_layout.addWidget(self.lbl_pending)
        stat_layout.addStretch(1)
        stat_layout.addWidget(self.lbl_progress)
        root.addLayout(stat_layout)

        out_group = QGroupBox("输出目录")
        out_layout = QHBoxLayout(out_group)
        self.output_dir_edit = QLineEdit()
        self.output_dir_edit.setPlaceholderText("选择表格后将自动设为表格所在目录")
        self.btn_pick_out = QPushButton("浏览…")
        self.btn_pick_out.clicked.connect(self._on_pick_output_dir)
        self.btn_open_out = QPushButton("打开输出文件夹")
        self.btn_open_out.clicked.connect(self._on_open_output_dir)
        out_layout.addWidget(self.output_dir_edit, 1)
        out_layout.addWidget(self.btn_pick_out)
        out_layout.addWidget(self.btn_open_out)
        root.addWidget(out_group)

        ctrl_layout = QHBoxLayout()
        self.btn_start_stop = QPushButton("开始")
        self.btn_start_stop.setEnabled(False)
        self.btn_start_stop.clicked.connect(self._on_start_stop)
        ctrl_layout.addWidget(self.btn_start_stop)
        ctrl_layout.addStretch(1)
        root.addLayout(ctrl_layout)

        log_header = QHBoxLayout()
        log_header.addWidget(QLabel("运行日志"))
        btn_clear_log = QPushButton("清空日志")
        btn_clear_log.clicked.connect(self._on_clear_log)
        log_header.addStretch(1)
        log_header.addWidget(btn_clear_log)
        root.addLayout(log_header)

        self.log_view = QPlainTextEdit()
        self.log_view.setReadOnly(True)
        self.log_view.setLineWrapMode(QPlainTextEdit.WidgetWidth)
        mono = QFont("Consolas")
        if not mono.exactMatch():
            mono = QFont("Courier New")
        self.log_view.setFont(mono)
        root.addWidget(self.log_view, 1)

        self.append_log("就绪。请选择 CSV 表格文件。")

    def append_log(self, message: str) -> None:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_view.appendPlainText(f"[{ts}] {message}")
        bar = self.log_view.verticalScrollBar()
        bar.setValue(bar.maximum())

    def closeEvent(self, event: QCloseEvent) -> None:
        if self._worker is not None and self._worker.isRunning():
            self.append_log("关闭窗口：请求停止任务…")
            if self._stop_event is not None:
                self._stop_event.set()
            self._worker.wait(5000)
        event.accept()

    def _set_busy(self, busy: bool) -> None:
        self._is_running = busy
        self.btn_pick_csv.setEnabled(not busy)
        self.btn_pick_out.setEnabled(not busy)
        self.btn_start_stop.setText("停止" if busy else "开始")
        self.btn_start_stop.setEnabled(busy or self._csv_path is not None)

    def _refresh_pending_count(self) -> None:
        self._total_rows = None
        self.lbl_pending.setText("待处理：— 条（商品ID 非空且含 SKU）")
        self.lbl_progress.setText("当前处理：—")
        if self._csv_path is None or not self._csv_path.is_file():
            return
        self.lbl_pending.setText("待处理：统计中…（商品ID 非空且含 SKU）")
        QApplication.processEvents()
        n, err = get_csv_data_row_count(self._csv_path)
        if n is not None:
            self._total_rows = n
            self.lbl_pending.setText(f"待处理：{n} 条（商品ID 非空且含 SKU）")
        else:
            self.lbl_pending.setText(f"待处理：无法统计（{err or '未知错误'}）")

    def _on_pick_csv(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择 CSV 表格",
            "",
            "CSV 文件 (*.csv);;所有文件 (*.*)",
        )
        if not path:
            return
        p = Path(path)
        if not p.is_file():
            QMessageBox.warning(self, "提示", "所选路径不是有效文件。")
            return
        self._csv_path = p.resolve()
        self.csv_path_edit.setText(str(self._csv_path))
        out_dir = str(self._csv_path.parent)
        self.output_dir_edit.setText(out_dir)
        self.append_log(f"已选择表格：{self._csv_path}")
        self.append_log(f"输出目录已设为：{out_dir}")
        self._refresh_pending_count()
        if not self._is_running:
            self.btn_start_stop.setEnabled(True)

    def _on_pick_output_dir(self) -> None:
        current = self.output_dir_edit.text().strip()
        start_dir = current if current and Path(current).is_dir() else str(Path.home())
        path = QFileDialog.getExistingDirectory(self, "选择输出目录", start_dir)
        if not path:
            return
        self.output_dir_edit.setText(path)
        self.append_log(f"输出目录已手动设为：{path}")

    def _on_open_output_dir(self) -> None:
        if self._last_material_root is not None and self._last_material_root.is_dir():
            p = self._last_material_root
        else:
            text = self.output_dir_edit.text().strip()
            if not text:
                QMessageBox.information(self, "提示", "请先设置输出目录。")
                return
            p = Path(text)
        if not open_local_dir(p):
            QMessageBox.warning(self, "提示", f"无法打开目录（路径不存在或无效）：\n{p}")

    def _on_worker_log(self, text: str) -> None:
        self.append_log(text)

    def _on_worker_progress(self, current: int, total: int) -> None:
        self.lbl_progress.setText(f"当前处理：{current}/{total}")

    def _on_material_root_created(self, material_root: str) -> None:
        self._last_material_root = Path(material_root)

    def _on_worker_finished(self, ok: bool, msg: str, material_root: str) -> None:
        self._worker = None
        self._stop_event = None
        self._set_busy(False)

        if ok and msg == "完成" and material_root:
            self._last_material_root = Path(material_root)
            total = self._total_rows if self._total_rows is not None else 0
            if total > 0:
                self.lbl_progress.setText(f"当前处理：{total}/{total}")
            else:
                self.lbl_progress.setText("当前处理：已完成")
            self.append_log(f"任务结束：{msg}")

            box = QMessageBox(self)
            box.setWindowTitle("完成")
            box.setIcon(QMessageBox.Information)
            box.setText("已完成所有处理。")
            btn_open = box.addButton("打开文件夹", QMessageBox.AcceptRole)
            box.addButton("取消", QMessageBox.RejectRole)
            box.setDefaultButton(btn_open)
            box.exec_()
            if box.clickedButton() == btn_open:
                rootp = Path(material_root)
                if not open_local_dir(rootp):
                    QMessageBox.warning(
                        self,
                        "提示",
                        f"无法打开素材包文件夹：\n{rootp}",
                    )
            self.lbl_progress.setText("当前处理：—")
        elif ok:
            self.append_log(f"任务结束：{msg}")
            self.lbl_progress.setText("当前处理：—")
        else:
            self.append_log(f"任务失败：{msg}")
            self.lbl_progress.setText("当前处理：—")
            QMessageBox.warning(self, "任务失败", msg)

    def _on_start_stop(self) -> None:
        if not self._is_running:
            if self._csv_path is None or not self._csv_path.is_file():
                QMessageBox.warning(self, "提示", "请先选择有效的 CSV 文件。")
                return
            out = self.output_dir_edit.text().strip()
            out_path = Path(out)
            if not out_path.is_dir():
                QMessageBox.warning(self, "提示", "请选择有效的输出目录。")
                return
            self.lbl_progress.setText("当前处理：准备中…")
            self._stop_event = threading.Event()
            self._last_material_root = None
            self._worker = PutawayThread(self._csv_path, out_path, self._stop_event)
            self._worker.log_message.connect(self._on_worker_log)
            self._worker.progress.connect(self._on_worker_progress)
            self._worker.material_root_created.connect(self._on_material_root_created)
            self._worker.finished.connect(self._on_worker_finished)
            self._set_busy(True)
            self.append_log("开始处理表格…")
            self._worker.start()
        else:
            if self._stop_event is not None:
                self._stop_event.set()
                self.append_log("已请求停止：当前文件写完后结束。")

    def _on_clear_log(self) -> None:
        self.log_view.clear()
        self.append_log("日志已清空。")


def main() -> None:
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
