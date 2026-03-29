from __future__ import annotations

import csv
import json
import re
import threading
from datetime import datetime
from pathlib import Path
from typing import Callable
from urllib.parse import urlparse

import requests
from charset_normalizer import from_bytes
from openpyxl import load_workbook

OnLog = Callable[[str], None]
OnProgress = Callable[[int, int], None]

CSV_PROBE_BYTES = 256 * 1024

FALLBACK_ENCODINGS: tuple[str, ...] = (
    "gb18030",
    "gbk",
    "cp936",
    "big5",
    "cp950",
    "utf-16-le",
    "utf-16-be",
    "utf-16",
)

PREFERRED_UTF: tuple[str, ...] = ("utf-8-sig", "utf-8")


def _normalize_enc_key(name: str) -> str:
    return name.replace("_", "-").lower()


def build_csv_encoding_candidates(sample: bytes) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()

    def add(enc: str | None) -> None:
        if not enc:
            return
        key = _normalize_enc_key(enc)
        if key in seen:
            return
        seen.add(key)
        ordered.append(enc)

    for enc in PREFERRED_UTF:
        add(enc)

    try:
        matches = from_bytes(sample)
        for i, m in enumerate(matches):
            if i >= 6:
                break
            add(m.encoding)
    except Exception:
        pass

    for enc in FALLBACK_ENCODINGS:
        add(enc)

    return ordered


def load_csv_with_encoding_detection(
    csv_path: Path,
    on_log: OnLog | None = None,
) -> tuple[list[str], list[dict[str, str]], str]:
    size = csv_path.stat().st_size
    if size == 0:
        raise ValueError("CSV 文件为空。")
    sample = csv_path.read_bytes()[: min(CSV_PROBE_BYTES, size)]
    candidates = build_csv_encoding_candidates(sample)
    if on_log:
        on_log(f"CSV 编码候选（探测 + 回退）: {', '.join(candidates[:8])}{'…' if len(candidates) > 8 else ''}")

    errors: list[str] = []
    for enc in candidates:
        try:
            with open(csv_path, "r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    errors.append(f"{enc}: 无表头")
                    continue
                headers = list(reader.fieldnames)
                rows = list(reader)
            return headers, rows, enc
        except UnicodeDecodeError as e:
            errors.append(f"{enc}: {e}")
        except UnicodeError as e:
            errors.append(f"{enc}: {e}")

    detail = "; ".join(errors[:5])
    if len(errors) > 5:
        detail += "…"
    raise ValueError(f"无法用已知编码解码 CSV。尝试摘要: {detail}")


WIN_ILLEGAL = r'<>:"/\\|?*'
WIN_ILLEGAL_SET = set(WIN_ILLEGAL)

SESSION_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "*/*",
}

IMG_SRC_RE = re.compile(
    r'<img[^>]+src\s*=\s*(["\'])(.*?)\1',
    re.IGNORECASE | re.DOTALL,
)

PIPE_DELIM = "'|'"

CONTENT_TYPE_EXT: dict[str, str] = {
    "image/jpeg": ".jpg",
    "image/jpg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "image/gif": ".gif",
    "image/bmp": ".bmp",
    "image/heic": ".heic",
    "image/heif": ".heif",
    "video/mp4": ".mp4",
    "video/quicktime": ".mov",
    "video/webm": ".webm",
    "video/x-msvideo": ".avi",
}

VIDEO_EXTS = {".mp4", ".mov", ".webm", ".avi", ".mkv", ".m4v", ".wmv"}
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp", ".heic", ".heif"}

MAX_PRODUCT_FOLDER_LEN = 120
MAX_STEM_SANITIZE = 200
MAX_SKU_NAME_STEM = 80


def sanitize_component(text: str, max_len: int) -> str:
    if not text:
        return ""
    s = "".join("_" if c in WIN_ILLEGAL_SET or ord(c) < 32 else c for c in str(text).strip())
    s = s.rstrip(" .")
    if not s:
        s = "_"
    if len(s) > max_len:
        s = s[:max_len].rstrip(" .") or "_"
    upper = s.upper()
    if upper in {"CON", "PRN", "AUX", "NUL", "COM1", "COM2", "COM3", "COM4", "LPT1", "LPT2", "LPT3"}:
        s = f"_{s}"
    return s


def product_folder_name(name: str, product_id: str) -> str:
    id_part = sanitize_component(product_id, 40)
    budget = MAX_PRODUCT_FOLDER_LEN - len(id_part) - 1
    if budget < 1:
        id_part = id_part[:20]
        budget = MAX_PRODUCT_FOLDER_LEN - len(id_part) - 1
    name_part = sanitize_component(name, min(budget, MAX_PRODUCT_FOLDER_LEN))
    out = f"{name_part}_{id_part}"
    if len(out) > MAX_PRODUCT_FOLDER_LEN:
        out = out[:MAX_PRODUCT_FOLDER_LEN].rstrip("._ ") or "_"
    return out


def root_material_folder_name(csv_stem: str) -> str:
    stem = sanitize_component(csv_stem, MAX_STEM_SANITIZE)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{stem}素材包{ts}"


def split_pipe_field(value: str | None) -> list[str]:
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    parts = [p.strip() for p in s.split(PIPE_DELIM)]
    return [p for p in parts if p]


def row_has_product_id(row: dict[str, str]) -> bool:
    return bool((row.get("商品ID") or "").strip())


def row_has_skus(row: dict[str, str]) -> bool:
    return len(split_pipe_field(row.get("SKU属性"))) > 0


def get_csv_data_row_count(csv_path: Path) -> tuple[int | None, str | None]:
    """统计有效数据行数：商品ID 非空且 SKU属性 至少有一段；不含表头。"""
    p = csv_path.resolve()
    if not p.is_file():
        return None, "文件不存在。"
    try:
        headers, rows, _enc = load_csv_with_encoding_detection(p, None)
    except ValueError as e:
        return None, str(e)
    if "商品ID" not in headers or "商品名称" not in headers:
        return None, "CSV 缺少必需列：商品ID 或 商品名称。"
    return sum(1 for r in rows if row_has_product_id(r) and row_has_skus(r)), None


PRODUCT_INFO_TEMPLATE = "商品信息.xlsx"

TPL_COL_TITLE = 1
TPL_COL_HUOHAO = 2
TPL_COL_ATTR = 3
TPL_COL_CATEGORY = 4
TPL_COL_BRAND = 5
TPL_COL_SPEC1 = 6
TPL_COL_PRICE = 8
TPL_COL_STOCK = 9
TPL_COL_SKU_BARCODE = 12


def product_info_template_path() -> Path:
    return Path(__file__).resolve().parent / PRODUCT_INFO_TEMPLATE


def format_product_attr_cell(raw: str | None) -> str:
    s = (raw or "").strip()
    if not s:
        return ""
    try:
        obj = json.loads(s)
    except json.JSONDecodeError:
        return s
    if not isinstance(obj, dict):
        return s
    parts = [f"{k}: {str(v)}" for k, v in obj.items()]
    return ", ".join(parts)


def _maybe_excel_number(s: str) -> str | int | float:
    t = (s or "").strip()
    if not t:
        return ""
    try:
        if "." in t:
            return float(t)
        return int(t)
    except ValueError:
        return t


def write_product_xlsx_from_template(template_path: Path, dest: Path, row: dict[str, str]) -> None:
    wb = load_workbook(template_path)
    ws = wb.active
    first_data_row = 3
    if ws.max_row >= first_data_row:
        ws.delete_rows(first_data_row, ws.max_row - first_data_row + 1)

    skus = split_pipe_field(row.get("SKU属性"))
    prices = split_pipe_field(row.get("SKU价格"))
    stocks = split_pipe_field(row.get("SKU库存"))
    barcodes = split_pipe_field(row.get("SKU条形码"))

    title = (row.get("商品名称") or "").strip()
    huohao = (row.get("商品ID") or "").strip()
    attr_text = format_product_attr_cell(row.get("商品属性"))
    category = (row.get("店铺分类") or "").strip()
    brand = "无"

    n = len(skus)
    for i in range(n):
        r = first_data_row + i
        spec = skus[i]
        price_s = prices[i] if i < len(prices) else ""
        stock_s = stocks[i] if i < len(stocks) else ""
        bc_s = barcodes[i] if i < len(barcodes) else ""

        if i == 0:
            ws.cell(row=r, column=TPL_COL_TITLE, value=title)
            ws.cell(row=r, column=TPL_COL_HUOHAO, value=huohao)
            ws.cell(row=r, column=TPL_COL_ATTR, value=attr_text)
            ws.cell(row=r, column=TPL_COL_CATEGORY, value=category)
            ws.cell(row=r, column=TPL_COL_BRAND, value=brand)
        else:
            for c in range(1, TPL_COL_BRAND + 1):
                ws.cell(row=r, column=c, value=None)

        ws.cell(row=r, column=TPL_COL_SPEC1, value=spec)
        ws.cell(row=r, column=TPL_COL_PRICE, value=_maybe_excel_number(price_s))
        ws.cell(row=r, column=TPL_COL_STOCK, value=_maybe_excel_number(stock_s))
        ws.cell(row=r, column=TPL_COL_SKU_BARCODE, value=bc_s or None)

    wb.save(dest)


def extension_from_url(url: str) -> str:
    try:
        suf = Path(urlparse(url).path).suffix.lower()
    except Exception:
        return ""
    if suf in IMAGE_EXTS or suf in VIDEO_EXTS:
        return suf
    if suf == ".jpe":
        return ".jpg"
    return ""


def extension_from_content_type(ct: str | None) -> str:
    if not ct:
        return ""
    base = ct.split(";")[0].strip().lower()
    return CONTENT_TYPE_EXT.get(base, "")


def pick_image_extension(url: str, content_type: str | None) -> str:
    u = extension_from_url(url)
    if u:
        return u
    c = extension_from_content_type(content_type)
    if c:
        return c
    return ".jpg"


def pick_video_extension(url: str, content_type: str | None) -> str:
    u = extension_from_url(url)
    if u and u in VIDEO_EXTS:
        return u
    c = extension_from_content_type(content_type)
    if c and c in VIDEO_EXTS:
        return c
    if u:
        return u
    return ".mp4"


def extract_img_srcs(html: str | None) -> list[str]:
    if not html:
        return []
    out: list[str] = []
    for m in IMG_SRC_RE.finditer(html):
        src = (m.group(2) or "").strip()
        if src.lower().startswith("http://") or src.lower().startswith("https://"):
            out.append(src)
    return out


def sanitize_file_stem(name: str, max_len: int = MAX_SKU_NAME_STEM) -> str:
    s = sanitize_component(name, max_len)
    s = s.replace(" ", "_")
    return s


def download_to_file(
    url: str,
    dest_without_ext: Path,
    stop_event: threading.Event,
    on_log: OnLog,
    *,
    is_video: bool,
) -> bool:
    if stop_event.is_set():
        return False
    last_err: str | None = None
    for attempt in range(1, 4):
        if stop_event.is_set():
            return False
        try:
            with requests.get(
                url,
                timeout=30,
                headers=SESSION_HEADERS,
                stream=True,
            ) as resp:
                resp.raise_for_status()
                ct = resp.headers.get("Content-Type")
                ext = (
                    pick_video_extension(url, ct)
                    if is_video
                    else pick_image_extension(url, ct)
                )
                final_path = dest_without_ext.with_suffix(ext)
                tmp_path = final_path.with_name(final_path.name + ".part")
                with open(tmp_path, "wb") as f:
                    for chunk in resp.iter_content(chunk_size=65536):
                        if chunk:
                            f.write(chunk)
                if final_path.exists():
                    final_path.unlink()
                tmp_path.replace(final_path)
            return True
        except Exception as e:
            last_err = str(e)
            on_log(f"下载失败（第 {attempt} 次）: {url[:80]}… — {e}")
    if last_err:
        on_log(f"已跳过（3 次均失败）: {url[:80]}…")
    return False


def run_job(
    csv_path: Path,
    output_base: Path,
    stop_event: threading.Event,
    on_log: OnLog,
    on_progress: OnProgress | None = None,
) -> tuple[bool, str, str]:
    csv_path = csv_path.resolve()
    output_base = output_base.resolve()
    if not csv_path.is_file():
        return False, "CSV 文件不存在。", ""
    if not output_base.is_dir():
        return False, "输出目录无效。", ""

    try:
        headers, rows, used_enc = load_csv_with_encoding_detection(csv_path, on_log)
    except ValueError as e:
        return False, str(e), ""

    on_log(f"已使用编码读取 CSV: {used_enc}")

    if "商品ID" not in headers or "商品名称" not in headers:
        return False, "CSV 缺少必需列：商品ID 或 商品名称。", ""

    tpl_path = product_info_template_path()
    if not tpl_path.is_file():
        return False, f"缺少商品信息模板文件：{tpl_path}", ""

    main_cols = [f"主图{i}" for i in range(1, 6)]

    root_name = root_material_folder_name(csv_path.stem)
    root_dir = output_base / root_name
    try:
        root_dir.mkdir(parents=True, exist_ok=False)
    except FileExistsError:
        return False, f"素材包目录已存在（可能同一秒内重复运行）: {root_dir}", ""
    except OSError as e:
        return False, f"无法创建素材包目录: {e}", ""

    material_root_str = str(root_dir)
    on_log(f"素材包目录: {root_dir}")

    raw_row_count = len(rows)
    rows_with_id = [r for r in rows if row_has_product_id(r)]
    skipped_empty_id = raw_row_count - len(rows_with_id)
    effective_rows = [r for r in rows_with_id if row_has_skus(r)]
    skipped_no_sku = len(rows_with_id) - len(effective_rows)
    total = len(effective_rows)
    parts_summary: list[str] = []
    if skipped_empty_id:
        parts_summary.append(f"{skipped_empty_id} 行商品ID 为空已忽略")
    if skipped_no_sku:
        parts_summary.append(f"{skipped_no_sku} 行 SKU属性 无有效分段已忽略")
    if parts_summary:
        on_log(
            f"CSV 共 {raw_row_count} 行数据，{', '.join(parts_summary)}，"
            f"待处理 {total} 条，开始处理。"
        )
    else:
        on_log(f"共 {total} 条有效商品记录，开始处理。")

    if total == 0:
        on_log("没有商品ID 非空的记录，无需处理。")
        on_log("全部商品处理完成。")
        return True, "完成", material_root_str

    for idx, row in enumerate(effective_rows, start=1):
        if stop_event.is_set():
            on_log("已收到停止指令，结束任务（当前文件已写完）。")
            return True, "已停止", material_root_str

        if on_progress is not None:
            on_progress(idx, total)

        name = (row.get("商品名称") or "").strip()
        pid = (row.get("商品ID") or "").strip()

        folder = product_folder_name(name or "_", pid)
        product_dir = root_dir / folder
        try:
            product_dir.mkdir(parents=False, exist_ok=False)
        except FileExistsError:
            on_log(f"[{idx}/{total}] 跳过：文件夹已存在 {folder}")
            continue
        except OSError as e:
            on_log(f"[{idx}/{total}] 跳过：无法创建文件夹 {folder} — {e}")
            continue

        sub_main = product_dir / "主图"
        sub_sku = product_dir / "SKU图"
        sub_detail = product_dir / "详情图"
        sub_video = product_dir / "主图视频"
        for d in (sub_main, sub_sku, sub_detail, sub_video):
            d.mkdir(parents=True, exist_ok=True)

        if stop_event.is_set():
            on_log("已收到停止指令，结束任务。")
            return True, "已停止", material_root_str

        xlsx_path = product_dir / "商品信息.xlsx"
        try:
            write_product_xlsx_from_template(tpl_path, xlsx_path, row)
            on_log(f"[{idx}/{total}] 已写入 商品信息.xlsx（{folder}）")
        except Exception as e:
            on_log(f"[{idx}/{total}] 写入 xlsx 失败: {e}")

        for col in main_cols:
            if col not in row:
                continue
            raw = (row.get(col) or "").strip()
            if not raw:
                continue
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str
            dest_stem = sub_main / col
            on_log(f"[{idx}/{total}] 下载 {col} …")
            ok = download_to_file(raw, dest_stem, stop_event, on_log, is_video=False)
            if ok:
                on_log(f"[{idx}/{total}] {col} 完成")
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str

        desc = row.get("商品描述图")
        imgs = extract_img_srcs(desc)
        for n, u in enumerate(imgs, start=1):
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str
            stem = sub_detail / f"详情图_{n}"
            on_log(f"[{idx}/{total}] 下载 详情图_{n} …")
            ok = download_to_file(u, stem, stop_event, on_log, is_video=False)
            if ok:
                on_log(f"[{idx}/{total}] 详情图_{n} 完成")
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str

        urls_sku = split_pipe_field(row.get("SKU属性图"))
        names_sku = split_pipe_field(row.get("SKU属性"))
        if len(urls_sku) != len(names_sku):
            on_log(
                f"[{idx}/{total}] 提示：SKU属性图（{len(urls_sku)} 个）与 "
                f"SKU属性（{len(names_sku)} 个）数量不一致，缺名将使用 SKU图_N。"
            )
        for si, u in enumerate(urls_sku, start=1):
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str
            if si <= len(names_sku) and names_sku[si - 1]:
                stem_name = sanitize_file_stem(names_sku[si - 1])
                if not stem_name:
                    stem_name = f"SKU图_{si}"
            else:
                stem_name = f"SKU图_{si}"
            dest_stem = sub_sku / stem_name
            on_log(f"[{idx}/{total}] 下载 SKU 图 ({stem_name}) …")
            ok = download_to_file(u, dest_stem, stop_event, on_log, is_video=False)
            if ok:
                on_log(f"[{idx}/{total}] SKU 图 {stem_name} 完成")
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str

        vid = (row.get("商品视频") or "").strip()
        if vid:
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str
            dest_stem = sub_video / "主图视频"
            on_log(f"[{idx}/{total}] 下载 商品视频 …")
            ok = download_to_file(vid, dest_stem, stop_event, on_log, is_video=True)
            if ok:
                on_log(f"[{idx}/{total}] 主图视频 完成")
            if stop_event.is_set():
                on_log("已收到停止指令，结束任务。")
                return True, "已停止", material_root_str

        on_log(f"[{idx}/{total}] 商品处理完成：{folder}")

    on_log("全部商品处理完成。")
    return True, "完成", material_root_str
